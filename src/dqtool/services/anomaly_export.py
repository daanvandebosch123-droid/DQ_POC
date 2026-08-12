from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="5B5248")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def build_anomaly_report_workbook(
    source_label: str, profile: dict[str, Any], anomalies: list[dict[str, Any]]
) -> bytes:
    """Create an Excel workbook for an anomaly profile without including source rows."""
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    _append_sheet(
        summary,
        ["Metric", "Value"],
        [
            ("Source", source_label),
            ("Profiled at (UTC)", profile.get("profiled_at") or ""),
            ("Rows", int(profile.get("row_count") or 0)),
            ("Columns profiled", len(profile.get("columns") or {})),
            ("Drift findings", len(anomalies)),
            ("High-severity findings", sum(1 for finding in anomalies if finding.get("severity") == "high")),
            ("GDPR review flags", len(profile.get("gdpr_findings") or [])),
        ],
    )

    findings = workbook.create_sheet("Drift findings")
    _append_sheet(
        findings,
        ["Severity", "Column", "Finding"],
        [
            (str(finding.get("severity") or "").upper(), finding.get("column") or "-", finding.get("message") or "")
            for finding in anomalies
        ],
    )

    columns = workbook.create_sheet("Column profile")
    _append_sheet(
        columns,
        ["Field", "Type", "Meaning", "SQL Null %", "Blank %", "Distinct", "Min", "Max", "Mean"],
        [
            (
                name,
                stats.get("type") or "",
                stats.get("inferred_type") or "text",
                _percent(stats.get("null_rate")),
                _percent(stats.get("blank_rate")),
                stats.get("distinct_count") if stats.get("distinct_count") is not None else "",
                stats.get("min") if stats.get("min") is not None else "",
                stats.get("max") if stats.get("max") is not None else "",
                stats.get("mean") if stats.get("mean") is not None else "",
            )
            for name, stats in profile.get("columns", {}).items()
        ],
    )

    frequencies = workbook.create_sheet("Value frequencies")
    _append_sheet(
        frequencies,
        ["Field", "Value", "Count", "Share"],
        [
            (name, item["value"], item["count"], item["share"])
            for name, stats in profile.get("columns", {}).items()
            for item in sorted(
                stats.get("frequency_values") or stats.get("top_values") or [],
                key=lambda item: (-int(item.get("count") or 0), str(item.get("value") or "")),
            )
        ],
    )

    gdpr = workbook.create_sheet("GDPR review")
    _append_sheet(
        gdpr,
        ["Severity", "Column", "Category", "Reason"],
        [
            (
                str(finding.get("severity") or "").upper(),
                finding.get("column") or "",
                finding.get("category") or "",
                finding.get("reason") or "",
            )
            for finding in profile.get("gdpr_findings") or []
        ],
    )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _append_sheet(sheet, headers: list[str], rows: list[tuple[Any, ...]]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(rows) + 1)}"
    for row in rows:
        sheet.append(row)
    for index, header in enumerate(headers, start=1):
        values = [str(header), *(str(row[index - 1] or "") for row in rows)]
        sheet.column_dimensions[get_column_letter(index)].width = min(60, max(12, max(map(len, values)) + 2))


def _percent(value: Any) -> float | str:
    return "" if value is None else float(value)
