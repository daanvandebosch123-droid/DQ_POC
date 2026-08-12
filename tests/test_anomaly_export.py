from __future__ import annotations

import unittest
from io import BytesIO

from openpyxl import load_workbook

from dqtool.services.anomaly_export import build_anomaly_report_workbook


class AnomalyExportTests(unittest.TestCase):
    def test_workbook_contains_anomaly_summary_findings_profile_and_gdpr_sheets(self) -> None:
        report = build_anomaly_report_workbook(
            "customers",
            {
                "profiled_at": "2026-08-12T10:00:00+00:00",
                "row_count": 12,
                "columns": {"email": {"type": "VARCHAR", "inferred_type": "email", "null_rate": 0.25}},
                "gdpr_findings": [
                    {"severity": "medium", "column": "email", "category": "Personal data", "reason": "Name signal"}
                ],
            },
            [{"severity": "high", "column": "email", "message": "Null rate increased"}],
        )

        workbook = load_workbook(BytesIO(report), data_only=True)

        self.assertEqual(["Summary", "Drift findings", "Column profile", "Value frequencies", "GDPR review"], workbook.sheetnames)
        self.assertEqual("customers", workbook["Summary"]["B2"].value)
        self.assertEqual("HIGH", workbook["Drift findings"]["A2"].value)
        self.assertEqual("email", workbook["Column profile"]["A2"].value)
        self.assertEqual("Personal data", workbook["GDPR review"]["C2"].value)

    def test_workbook_exports_every_available_frequency_value(self) -> None:
        report = build_anomaly_report_workbook(
            "customers",
            {
                "row_count": 10,
                "columns": {
                    "status": {
                        "frequency_values": [
                            {"value": "inactive", "count": 2, "share": 0.2},
                            {"value": "active", "count": 8, "share": 0.8},
                        ],
                        "top_values": [{"value": "active", "count": 8, "share": 0.8}],
                    }
                },
            },
            [],
        )

        workbook = load_workbook(BytesIO(report), data_only=True)
        rows = list(workbook["Value frequencies"].values)

        self.assertEqual(("status", "active", 8, 0.8), rows[1])
        self.assertEqual(("status", "inactive", 2, 0.2), rows[2])
